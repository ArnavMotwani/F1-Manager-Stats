from DriverEntry import driver
from openpyxl import load_workbook, Workbook
from Cells import outputs, inputs

data = load_workbook(filename="FOneData.xlsx")
Names, Levels, Length = driver()
sheet = data["Driver Info"]

file = Workbook()
output = file.active

for i in range(Length):
    for row in range(1, 168):
        cellN, cellL, cellO, cellD, cellC, cellF, cellT, cellW = inputs(row)
        valueN = sheet[cellN].value
        valueL = sheet[cellL].value
        valueO = sheet[cellO].value
        valueD = sheet[cellD].value
        valueC = sheet[cellC].value
        valueF = sheet[cellF].value
        valueT = sheet[cellT].value
        valueW = sheet[cellW].value

        if valueN == Names[i]:
            if valueL == Levels[i]:
                name, lvl, ot, df, cn, fm, tm, wm = outputs(i)
                output[name] = valueN
                output[lvl] = valueL
                output[ot] = valueO
                output[df] = valueD
                output[cn] = valueC
                output[fm] = valueF
                output[tm] = valueT
                output[wm] = valueW

output["A1"] = "Driver"
output["B1"] = "Level"
output["C1"] = "Overtaking"
output["D1"] = "Defending"
output["E1"] = "Consistency"
output["F1"] = "Fuel mgt"
output["G1"] = "Tire mgt"
output["H1"] = "Wet mgt"

file.save("output.xlsx")
